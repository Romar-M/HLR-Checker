"""
Модуль работы с Excel-файлами для HLR Checker.

Функции:
- загрузка книги;
- удаление пустых столбцов (заголовок есть, а данных нет);
- автоматический поиск столбцов с телефонными номерами;
- нормализация номеров прямо в исходных ячейках (формат 8XXXXXXXXXX);
- проверка номеров (реальный HLR или демо);
- добавление справа столбцов: <колонка>_статус, <колонка>_оператор,
  <колонка>_регион и итогового "проверка";
- сохранение результата.

ВАЖНО: исходные данные сохраняются, но пустые столбцы удаляются,
чтобы обработанный файл был компактным и его было удобно загружать обратно
(например, в amoCRM).
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from hlr_checker import check_numbers, normalize_phone


# Ключевые слова для поиска столбцов с телефонами в заголовках.
PHONE_COLUMN_KEYWORDS = [
    "телефон",
    "phone",
    "мобильный",
    "mob",
    "сотовый",
    "сот",
    "номер",
    "тел",
    "контактный телефон",
]


def find_phone_columns(ws, max_scan_rows=50):
    """
    Ищет столбцы, в которых лежат телефонные номера.

    1) По заголовкам (первая строка), но только если содержимое колонки
       действительно похоже на телефоны (проверка по данным).
    2) Если заголовков нет — по содержимому первых строк.

    Возвращает список индексов колонок (1-based, как в openpyxl).
    """
    # 1) Кандидаты по заголовкам.
    candidates = []
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if header is None:
            continue
        header_lower = str(header).lower()
        if any(kw in header_lower for kw in PHONE_COLUMN_KEYWORDS):
            candidates.append(c)

    # Проверяем кандидатов по фактическим данным (пропускаем строку заголовка).
    verified = [
        c for c in candidates
        if _column_has_phones(ws, c, start_row=2, max_scan_rows=max_scan_rows)
    ]
    if verified:
        return verified

    # 2) Резервный поиск по содержимому (когда заголовков нет).
    phone_cols = []
    for c in range(1, ws.max_column + 1):
        if _column_has_phones(ws, c, start_row=1, max_scan_rows=max_scan_rows):
            phone_cols.append(c)

    return phone_cols


def _column_has_phones(ws, col_idx, start_row=1, max_scan_rows=50):
    """
    Возвращает True, если большинство непустых значений колонки — телефоны.
    """
    total = 0
    matched = 0
    last_row = min(start_row + max_scan_rows - 1, ws.max_row)
    for r in range(start_row, last_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val is None or str(val).strip() == "":
            continue
        total += 1
        if normalize_phone(val) is not None:
            matched += 1

    return total > 0 and matched / total >= 0.6


def _is_non_empty(value):
    """True, если значение ячейки не является пустым."""
    return value is not None and str(value).strip() != ""


def remove_empty_columns(ws, data_start_row=2):
    """
    Удаляет столбцы, в которых нет ни одного непустого значения
    в строках данных (начиная с data_start_row).

    Заголовок (первая строка) при этом НЕ считается данными: если у столбца
    есть заголовок, но под ним нет ни одного значения — столбец считается
    пустым и удаляется.

    Удаление выполняется справа налево, чтобы индексы колонок не смещались.
    Возвращает список удалённых индексов (1-based, до удаления).
    """
    if ws.max_row < data_start_row:
        return []

    empty_cols = []
    for c in range(1, ws.max_column + 1):
        has_data = any(
            _is_non_empty(ws.cell(row=r, column=c).value)
            for r in range(data_start_row, ws.max_row + 1)
        )
        if not has_data:
            empty_cols.append(c)

    for c in reversed(empty_cols):
        ws.delete_cols(c, 1)

    return empty_cols


def process_workbook(input_path, output_path, api_key=""):
    """
    Основная функция обработки файла.

    1. Загружает Excel.
    2. Удаляет пустые столбцы (нет данных под заголовком).
    3. Для каждого листа находит столбцы с телефонами.
    4. Нормализует номера прямо в исходных ячейках (к 8XXXXXXXXXX).
    5. Проверяет номера (реальный HLR, если задан api_key, иначе демо).
    6. Добавляет справа столбцы:
         <имя>_статус, <имя>_оператор, <имя>_регион
       и итоговый "проверка".
    7. Сохраняет результат в output_path.

    Возвращает общее количество удалённых пустых столбцов.
    """
    wb = load_workbook(input_path)
    found_any = False
    total_removed_cols = 0

    for ws in wb.worksheets:
        # Определяем, с какой строки начинаются данные.
        prelim_phone_cols = find_phone_columns(ws)
        data_start_row = 2 if _row_is_header(ws, prelim_phone_cols) else 1

        # Удаляем столбцы, в которых нет данных.
        removed = remove_empty_columns(ws, data_start_row)
        total_removed_cols += len(removed)

        # Находим телефонные столбцы заново: после удаления индексы сдвинулись.
        phone_cols = find_phone_columns(ws)
        if not phone_cols:
            continue

        found_any = True

        first_data_row = 2 if _row_is_header(ws, phone_cols) else 1

        # Обрабатываем каждый телефонный столбец.
        per_column = []
        for col_idx in phone_cols:
            col_header = ws.cell(row=1, column=col_idx).value
            col_name = str(col_header).strip() if col_header else f"колонка_{col_idx}"

            statuses = []
            results = []
            operators = []
            regions = []

            for r in range(first_data_row, ws.max_row + 1):
                raw = ws.cell(row=r, column=col_idx).value
                res = check_numbers([raw], api_key)[0]

                if api_key:
                    print(
                        f"      [HLR]  {res['phone']} -> {res['status']} "
                        f"({res.get('operator', '—')}, {res.get('region', '—')})",
                        flush=True,
                    )

                # Переписываем исходную ячейку в формат 8XXXXXXXXXX.
                # Для некорректных значений оставляем исходный текст.
                if res["result"] != "invalid":
                    ws.cell(row=r, column=col_idx, value=res["phone"])

                statuses.append(res["status"])
                results.append(res["result"])
                operators.append(res.get("operator", "—"))
                regions.append(res.get("region", "—"))

            per_column.append({
                "col_idx": col_idx,
                "col_name": col_name,
                "statuses": statuses,
                "results": results,
                "operators": operators,
                "regions": regions,
            })

        # Запоминаем, с какой колонки начинаются добавляемые столбцы.
        first_new_col = ws.max_column + 1

        # Добавляем справа новые столбцы.
        added_status_cols = {}
        for info in per_column:
            status_col = ws.max_column + 1
            operator_col = status_col + 1
            region_col = status_col + 2

            ws.cell(row=1, column=status_col, value=f"{info['col_name']}_статус")
            ws.cell(row=1, column=operator_col, value=f"{info['col_name']}_оператор")
            ws.cell(row=1, column=region_col, value=f"{info['col_name']}_регион")

            for i in range(len(info["statuses"])):
                r = first_data_row + i
                ws.cell(row=r, column=status_col, value=info["statuses"][i])
                ws.cell(row=r, column=operator_col, value=info["operators"][i])
                ws.cell(row=r, column=region_col, value=info["regions"][i])

            added_status_cols[info["col_idx"]] = status_col

        # Итоговая колонка "проверка".
        result_col = ws.max_column + 1
        ws.cell(row=1, column=result_col, value="проверка")

        for i, r in enumerate(range(first_data_row, ws.max_row + 1)):
            row_results = [info["results"][i] for info in per_column]

            if "ok" in row_results:
                summary = "активен"
                color = "008000"
            elif "inactive" in row_results:
                summary = "заблокирован"
                color = "CC0000"
            elif "unreachable" in row_results:
                summary = "недоступен"
                color = "E67E22"
            else:
                summary = "ошибка формата"
                color = "808080"

            cell = ws.cell(row=r, column=result_col, value=summary)
            cell.font = Font(color=color, bold=True)

        # Оформляем шапки ТОЛЬКО добавленных колонок (исходные не трогаем).
        _style_new_headers(ws, first_new_col)

    if not found_any:
        raise ValueError(
            "Не удалось найти столбцы с телефонными номерами. "
            "Убедитесь, что в файле есть колонки с номерами телефонов."
        )

    wb.save(output_path)
    return total_removed_cols


def _row_is_header(ws, phone_cols):
    """
    Определяет, является ли первая строка заголовком.

    Если в первой строке в найденных телефонных столбцах лежат
    НЕ-номера (текст), считаем её заголовком.
    """
    for col_idx in phone_cols:
        val = ws.cell(row=1, column=col_idx).value
        if val is not None and normalize_phone(val) is None:
            return True
    return False


def _style_new_headers(ws, first_new_col):
    """Оформляет шапки только добавленных (справа) колонок."""
    for c in range(first_new_col, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

